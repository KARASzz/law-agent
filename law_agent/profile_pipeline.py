# -*- coding: utf-8 -*-
"""
画像数据批处理管线。

这个模块承接原前置工具的 Excel 清洗逻辑，并补充助理协作记录的候选画像池。
所有能力都面向命令行/批处理入口，不在 Web 工作台暴露入口。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from law_agent.client_profiles import ClientProfileStore


# ============================================================
# 1. 正式画像采集表字段映射
# ============================================================

HEADER_SYNONYMS: Dict[str, List[str]] = {
    "collection_date": ["采集日期", "采集日期*"],
    "collection_id": ["采集ID", "脱敏案号"],
    "data_source": ["数据来源", "数据来源*"],
    "matter_type": ["案件类型", "案件类型*", "业务类型"],
    "stage": ["阶段", "阶段*", "案件阶段"],
    "conflict_structure": ["冲突结构", "冲突结构*"],
    "role_pattern": ["角色格局"],
    "client_goal": ["客户核心诉求"],
    "key_constraints": ["关键约束"],
    "first_judgment": ["第一判断", "第一判断*"],
    "abstract_reason": ["判断理由_抽象", "判断理由"],
    "strategy_choice": ["策略选择", "策略选择*"],
    "value_order": ["价值排序"],
    "risk_communication": ["风险沟通方式"],
    "handling_temperature": ["处理温度"],
    "result_review": ["结果/复盘", "复盘一句话_可空"],
    "reusable_rule": ["可复用规则"],
    "representativeness": ["代表性"],
    "note": ["备注_不写细节", "备注"],
    "desensitization_status": ["脱敏状态", "脱敏状态*"],
    "ingestion_level": ["入库等级", "入库等级*"],
    "cleaning_note": ["清洗备注"],
    "review_tag": ["复盘标签"],
    "profile_update_action": ["画像更新动作"],
}

CORE_FIELDS = [
    "conflict_structure",
    "first_judgment",
    "strategy_choice",
    "reusable_rule",
]

OUTPUT_FIELD_ORDER = [
    "collection_date",
    "collection_id",
    "data_source",
    "matter_type",
    "stage",
    "conflict_structure",
    "role_pattern",
    "client_goal",
    "key_constraints",
    "first_judgment",
    "abstract_reason",
    "strategy_choice",
    "value_order",
    "risk_communication",
    "handling_temperature",
    "result_review",
    "reusable_rule",
    "representativeness",
    "note",
    "desensitization_status",
    "ingestion_level",
    "cleaning_note",
    "review_tag",
    "profile_update_action",
]


# ============================================================
# 2. 助理协作记录字段映射
# ============================================================

ASSISTANT_HEADER_SYNONYMS: Dict[str, List[str]] = {
    "work_date": ["日期"],
    "assistant_name": ["助理姓名"],
    "lawyer_name": ["服务律师"],
    "matter_code": ["事项代号"],
    "work_type": ["工作类型"],
    "task_source": ["任务来源"],
    "assigned_task": ["律师交代的任务"],
    "completed_work": ["我完成了什么"],
    "deliverable_type": ["交付物类型"],
    "duration_minutes": ["用时_分钟", "用时分钟"],
    "current_status": ["当前状态"],
    "blocker_reason": ["卡点原因"],
    "lawyer_feedback": ["律师反馈或修改方向"],
    "next_follow_up": ["下步跟进"],
    "habit_suitability": ["是否适合沉淀为习惯"],
    "note": ["备注"],
}

ASSISTANT_FIELD_ORDER = [
    "work_date",
    "assistant_name",
    "lawyer_name",
    "matter_code",
    "work_type",
    "task_source",
    "assigned_task",
    "completed_work",
    "deliverable_type",
    "duration_minutes",
    "current_status",
    "blocker_reason",
    "lawyer_feedback",
    "next_follow_up",
    "habit_suitability",
    "note",
]

ASSISTANT_REQUIRED_HEADERS = [
    "work_date",
    "assistant_name",
    "lawyer_name",
    "work_type",
    "lawyer_feedback",
    "habit_suitability",
]


# ============================================================
# 3. 脱敏与通用工具
# ============================================================

SENSITIVE_PATTERNS: List[Tuple[str, re.Pattern[str]]] = [
    ("身份证号", re.compile(r"(?<!\d)\d{17}[\dXx](?!\d)")),
    ("手机号", re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)")),
    ("邮箱", re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")),
    ("疑似案号", re.compile(r"[（(]\d{4}[）)][^，。；;\s]{2,40}?号")),
    ("疑似金额", re.compile(r"\d+(?:\.\d+)?\s*(?:万|万元|元|亿|亿元)")),
    ("疑似具体日期", re.compile(r"\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}日?")),
]

SENSITIVE_FLAG_ONLY: List[Tuple[str, re.Pattern[str]]] = [
    ("疑似具体司法/行政机构", re.compile(r"[^，。；;\s]{2,20}(人民法院|检察院|公安局|仲裁委员会|市场监督管理局)")),
    ("疑似自然人姓名提示", re.compile(r"(张三|李四|王五|某某|当事人姓名|对方姓名|法官姓名|律师姓名)")),
]


def normalize_header(text: Any) -> str:
    if text is None:
        return ""
    return str(text).strip().replace("*", "").replace(" ", "")


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def sanitize_text(value: Any) -> Tuple[str, List[str]]:
    """返回脱敏后的文本和命中的敏感风险标签。"""
    text = stringify_cell(value)
    if not text:
        return "", []

    flags: List[str] = []
    sanitized = text

    for name, pattern in SENSITIVE_PATTERNS:
        if pattern.search(sanitized):
            flags.append(name)
            sanitized = pattern.sub(f"[已脱敏_{name}]", sanitized)

    for name, pattern in SENSITIVE_FLAG_ONLY:
        if pattern.search(sanitized):
            flags.append(name)

    if len(sanitized) > 800:
        flags.append("文本过长_建议人工复核")
        sanitized = sanitized[:800] + "...[已截断_待复核]"

    return sanitized, sorted(set(flags))


def parse_bool(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "启用"}


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def stable_record_id(source_file: str, sheet_name: str, row_number: int, row_payload: Dict[str, str]) -> str:
    seed = json.dumps(
        {
            "source_file": source_file,
            "sheet_name": sheet_name,
            "row_number": row_number,
            "core": {key: row_payload.get(key, "") for key in CORE_FIELDS},
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    return hashlib.sha256(seed.encode("utf-8")).hexdigest()[:20]


def stable_candidate_id(source_file: str, sheet_name: str, row_number: int, row_payload: Dict[str, str]) -> str:
    seed = json.dumps(
        {
            "source_file": source_file,
            "sheet_name": sheet_name,
            "row_number": row_number,
            "core": {
                "work_type": row_payload.get("work_type", ""),
                "lawyer_feedback": row_payload.get("lawyer_feedback", ""),
                "habit_suitability": row_payload.get("habit_suitability", ""),
            },
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    return "cand_" + hashlib.sha256(seed.encode("utf-8")).hexdigest()[:20]


def stable_promoted_record_id(candidate_id: str, candidate_rule: str) -> str:
    seed = json.dumps(
        {
            "candidate_id": candidate_id,
            "candidate_rule": candidate_rule,
            "kind": "assistant_profile_candidate_promotion",
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    return hashlib.sha256(seed.encode("utf-8")).hexdigest()[:20]


def _safe_output_name(stem: str) -> str:
    return re.sub(r"[^0-9A-Za-z_\-\u4e00-\u9fa5]", "_", stem)


def _write_json(payload: Dict[str, Any], output_path: Path, pretty: bool = True) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2 if pretty else None)


def _build_synonym_lookup(synonyms: Dict[str, List[str]]) -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    for canonical, aliases in synonyms.items():
        for alias in aliases:
            lookup[normalize_header(alias)] = canonical
    return lookup


def _find_header_row(
    ws,
    synonyms: Dict[str, List[str]],
    max_scan_rows: int = 20,
    min_score: int = 4,
    required_fields: Optional[Sequence[str]] = None,
) -> Tuple[int, Dict[str, int]]:
    lookup = _build_synonym_lookup(synonyms)
    best_row = -1
    best_score = 0
    best_map: Dict[str, int] = {}

    for row_idx in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_map: Dict[str, int] = {}
        score = 0
        for col_idx in range(1, ws.max_column + 1):
            header = normalize_header(ws.cell(row_idx, col_idx).value)
            if header in lookup:
                canonical = lookup[header]
                row_map.setdefault(canonical, col_idx)
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_map = row_map

    missing_required = [
        field for field in (required_fields or [])
        if field not in best_map
    ]
    if best_score < min_score or missing_required:
        if missing_required:
            raise ValueError(f"未识别到必要表头：{missing_required}")
        raise ValueError("未识别到有效表头。请确认工作表表头没有被大幅改名。")

    return best_row, best_map


def choose_sheet(wb, requested_sheet: Optional[str], sheet_candidates: List[str]) -> str:
    if requested_sheet:
        if requested_sheet not in wb.sheetnames:
            raise ValueError(f"指定工作表不存在：{requested_sheet}；当前工作表：{wb.sheetnames}")
        return requested_sheet

    for name in sheet_candidates:
        if name in wb.sheetnames:
            return name

    for name in wb.sheetnames:
        if "填写表" in name or "采集表" in name:
            return name

    raise ValueError(f"没有找到采集表工作表。当前工作表：{wb.sheetnames}")


def choose_assistant_sheet(wb, requested_sheet: Optional[str], sheet_candidates: List[str]) -> str:
    if requested_sheet:
        if requested_sheet not in wb.sheetnames:
            raise ValueError(f"指定工作表不存在：{requested_sheet}；当前工作表：{wb.sheetnames}")
        return requested_sheet

    for name in sheet_candidates:
        if name in wb.sheetnames:
            return name

    for name in wb.sheetnames:
        if "协作记录" in name:
            return name

    raise ValueError(f"没有找到助理协作记录工作表。当前工作表：{wb.sheetnames}")


def _resolve_config_path_value(value: str, config_dir: Optional[Path]) -> str:
    if not value:
        return value
    path = Path(value)
    if path.is_absolute() or config_dir is None:
        return str(path)
    return str((config_dir / path).resolve())


def load_pipeline_config(path: Optional[Path]) -> Dict[str, Any]:
    default = {
        "schema_version": "lawyer_profile_memory_v1",
        "sheet_candidates": ["01_用户填写表", "01_极简采集表"],
        "candidate_db_path": "data/profile_candidates.db",
        "client_profile_db_path": "data/client_profiles.db",
        "output": {"pretty_json": True, "also_write_jsonl": False},
        "assistant_candidates": {
            "schema_version": "lawyer_profile_candidate_v1",
            "sheet_candidates": ["每日协作记录"],
            "include_uncertain": False,
            "default_ingestion_level": "B可用样本",
        },
        "model": {"use_model": False},
    }

    config_dir: Optional[Path] = None
    if path and path.exists():
        config_dir = path.resolve().parent
        with path.open("r", encoding="utf-8") as f:
            loaded = json.load(f)
        for key, value in loaded.items():
            if isinstance(value, dict) and isinstance(default.get(key), dict):
                default[key].update(value)
            else:
                default[key] = value

    default["candidate_db_path"] = _resolve_config_path_value(
        str(default.get("candidate_db_path", "")),
        config_dir,
    )
    default["client_profile_db_path"] = _resolve_config_path_value(
        str(default.get("client_profile_db_path", "")),
        config_dir,
    )
    return default


# ============================================================
# 4. 模型占位符
# ============================================================

def normalize_with_model(record: Dict[str, Any], model_config: Dict[str, Any]) -> Dict[str, Any]:
    """
    可选：用模型进一步把采集记录归一化为画像规则。

    默认不调用外部模型。当前项目默认保持 use_model=false。
    """
    record["model_normalized"] = None
    record["model_note"] = "model disabled; fill normalize_with_model() when needed"
    return record


# ============================================================
# 5. 正式画像清洗
# ============================================================

def clean_profile_workbook(
    input_path: Path,
    output_dir: Path,
    config: Dict[str, Any],
    requested_sheet: Optional[str] = None,
    use_model: bool = False,
) -> Path:
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, data_only=True, read_only=True)
    sheet_name = choose_sheet(wb, requested_sheet, config.get("sheet_candidates", []))
    ws = wb[sheet_name]

    header_row, header_map = _find_header_row(ws, HEADER_SYNONYMS)

    records: List[Dict[str, Any]] = []
    stats = {
        "total_excel_rows_scanned": 0,
        "empty_rows_skipped": 0,
        "records_exported": 0,
        "records_need_manual_review": 0,
        "sensitive_flag_counts": {},
        "missing_core_field_counts": {field: 0 for field in CORE_FIELDS},
    }

    source_hash = file_sha256(input_path)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        stats["total_excel_rows_scanned"] += 1

        raw_payload: Dict[str, str] = {}
        sensitive_flags_by_field: Dict[str, List[str]] = {}
        all_flags: List[str] = []

        for field in OUTPUT_FIELD_ORDER:
            col_idx = header_map.get(field)
            raw_value = ws.cell(row_idx, col_idx).value if col_idx else ""
            if field == "collection_date":
                sanitized, flags = stringify_cell(raw_value), []
            else:
                sanitized, flags = sanitize_text(raw_value)
            raw_payload[field] = sanitized
            if flags:
                sensitive_flags_by_field[field] = flags
                all_flags.extend(flags)

        if not any(raw_payload.values()):
            stats["empty_rows_skipped"] += 1
            continue

        missing_core = [field for field in CORE_FIELDS if not raw_payload.get(field)]
        for field in missing_core:
            stats["missing_core_field_counts"][field] += 1

        all_flags_unique = sorted(set(all_flags))
        for flag in all_flags_unique:
            stats["sensitive_flag_counts"][flag] = stats["sensitive_flag_counts"].get(flag, 0) + 1

        privacy_decision = "pass"
        if all_flags_unique:
            privacy_decision = "needs_manual_review"
        if len(missing_core) >= 3:
            privacy_decision = "low_value_or_incomplete"
        if privacy_decision == "needs_manual_review":
            stats["records_need_manual_review"] += 1

        record = {
            "record_id": stable_record_id(input_path.name, sheet_name, row_idx, raw_payload),
            "source": {
                "source_file": input_path.name,
                "source_file_sha256": source_hash,
                "sheet_name": sheet_name,
                "excel_row_number": row_idx,
            },
            "taxonomy": {
                "data_source": raw_payload.get("data_source", ""),
                "matter_type": raw_payload.get("matter_type", ""),
                "stage": raw_payload.get("stage", ""),
                "representativeness": raw_payload.get("representativeness", ""),
            },
            "judgment_model": {
                "conflict_structure": raw_payload.get("conflict_structure", ""),
                "role_pattern": raw_payload.get("role_pattern", ""),
                "client_goal": raw_payload.get("client_goal", ""),
                "key_constraints": raw_payload.get("key_constraints", ""),
                "first_judgment": raw_payload.get("first_judgment", ""),
                "abstract_reason": raw_payload.get("abstract_reason", ""),
                "strategy_choice": raw_payload.get("strategy_choice", ""),
                "value_order": raw_payload.get("value_order", ""),
                "risk_communication": raw_payload.get("risk_communication", ""),
                "handling_temperature": raw_payload.get("handling_temperature", ""),
                "reusable_rule": raw_payload.get("reusable_rule", ""),
            },
            "review_and_ingestion": {
                "collection_date": raw_payload.get("collection_date", ""),
                "collection_id_or_masked_ref": raw_payload.get("collection_id", ""),
                "result_review": raw_payload.get("result_review", ""),
                "note": raw_payload.get("note", ""),
                "desensitization_status": raw_payload.get("desensitization_status", ""),
                "ingestion_level": raw_payload.get("ingestion_level", ""),
                "cleaning_note": raw_payload.get("cleaning_note", ""),
                "review_tag": raw_payload.get("review_tag", ""),
                "profile_update_action": raw_payload.get("profile_update_action", ""),
            },
            "quality_control": {
                "missing_core_fields": missing_core,
                "privacy_decision": privacy_decision,
                "sensitive_flags": all_flags_unique,
                "sensitive_flags_by_field": sensitive_flags_by_field,
            },
        }

        if use_model:
            record = normalize_with_model(record, config.get("model", {}))

        records.append(record)

    stats["records_exported"] = len(records)

    exported = {
        "meta": {
            "schema_version": config.get("schema_version", "lawyer_profile_memory_v1"),
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "input_file": input_path.name,
            "input_file_sha256": source_hash,
            "sheet_name": sheet_name,
            "header_row": header_row,
            "privacy_principle": "仅输出脱敏后的抽象行为逻辑字段；不输出整行原始案件信息。",
        },
        "stats": stats,
        "records": records,
    }

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{_safe_output_name(input_path.stem)}_cleaned_{timestamp}.json"

    pretty = bool(config.get("output", {}).get("pretty_json", True))
    _write_json(exported, output_path, pretty=pretty)

    if bool(config.get("output", {}).get("also_write_jsonl", False)):
        jsonl_path = output_dir / f"{_safe_output_name(input_path.stem)}_records_{timestamp}.jsonl"
        with jsonl_path.open("w", encoding="utf-8") as f:
            for record in records:
                f.write(json.dumps(record, ensure_ascii=False) + "\n")

    return output_path


# ============================================================
# 6. 助理协作候选画像池
# ============================================================

@dataclass
class CandidateIngestionSummary:
    batch_id: str
    source_file: str
    candidates_seen: int
    candidates_upserted: int
    db_path: str


@dataclass
class CandidatePromotionSummary:
    output_file: str
    records_seen: int
    records_promoted: int
    records_skipped: int
    import_id: str
    client_profile_db_path: str


class ProfileCandidateStore:
    """SQLite 版助理协作候选画像池。"""

    def __init__(self, db_path: str = "data/profile_candidates.db"):
        self.db_path = db_path
        self._init_database()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_database(self) -> None:
        Path(self.db_path).parent.mkdir(parents=True, exist_ok=True)
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS profile_candidate_batches (
                    batch_id TEXT PRIMARY KEY,
                    schema_version TEXT,
                    generated_at TEXT,
                    input_file TEXT,
                    input_file_sha256 TEXT,
                    json_file_path TEXT NOT NULL,
                    json_file_sha256 TEXT NOT NULL,
                    stats_json TEXT NOT NULL,
                    created_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS profile_candidates (
                    candidate_id TEXT PRIMARY KEY,
                    latest_batch_id TEXT NOT NULL,
                    source_file TEXT,
                    source_file_sha256 TEXT,
                    sheet_name TEXT,
                    excel_row_number INTEGER,
                    work_date TEXT,
                    assistant_name TEXT,
                    lawyer_name TEXT,
                    matter_code TEXT,
                    work_type TEXT,
                    task_source TEXT,
                    assigned_task TEXT,
                    completed_work TEXT,
                    deliverable_type TEXT,
                    duration_minutes TEXT,
                    current_status TEXT,
                    blocker_reason TEXT,
                    lawyer_feedback TEXT,
                    next_follow_up TEXT,
                    habit_suitability TEXT,
                    note TEXT,
                    candidate_rule TEXT,
                    quality_decision TEXT,
                    sensitive_flags_json TEXT NOT NULL,
                    sensitive_flags_by_field_json TEXT NOT NULL,
                    candidate_json TEXT NOT NULL,
                    review_status TEXT NOT NULL,
                    promotion_status TEXT NOT NULL,
                    promoted_record_id TEXT,
                    promoted_import_id TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY(latest_batch_id) REFERENCES profile_candidate_batches(batch_id)
                )
                """
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_candidates_quality ON profile_candidates(quality_decision)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_candidates_promotion ON profile_candidates(promotion_status)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_candidates_review ON profile_candidates(review_status)"
            )

    def ingest_candidate_json_file(self, json_file_path: str) -> CandidateIngestionSummary:
        path = Path(json_file_path)
        payload = json.loads(path.read_text(encoding="utf-8"))
        meta = payload.get("meta", {})
        stats = payload.get("stats", {})
        candidates = payload.get("candidates", [])

        batch_id = self._build_batch_id(meta, path)
        created_at = datetime.now().isoformat(timespec="seconds")
        json_sha256 = file_sha256(path)

        upserted = 0
        with self._connect() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO profile_candidate_batches (
                    batch_id, schema_version, generated_at, input_file,
                    input_file_sha256, json_file_path, json_file_sha256,
                    stats_json, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    meta.get("schema_version"),
                    meta.get("generated_at"),
                    meta.get("input_file"),
                    meta.get("input_file_sha256"),
                    str(path),
                    json_sha256,
                    json.dumps(stats, ensure_ascii=False),
                    created_at,
                ),
            )

            for candidate in candidates:
                self._upsert_candidate(conn, candidate, batch_id, created_at)
                upserted += 1

        return CandidateIngestionSummary(
            batch_id=batch_id,
            source_file=str(path),
            candidates_seen=len(candidates),
            candidates_upserted=upserted,
            db_path=self.db_path,
        )

    def list_candidates(
        self,
        review_status: Optional[str] = None,
        promotion_status: Optional[str] = None,
        quality_decision: Optional[str] = None,
        limit: int = 20,
    ) -> List[Dict[str, Any]]:
        conditions: List[str] = []
        params: List[Any] = []
        filters = {
            "review_status": review_status,
            "promotion_status": promotion_status,
            "quality_decision": quality_decision,
        }
        for column, value in filters.items():
            if value:
                conditions.append(f"{column} = ?")
                params.append(value)

        where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        params.append(limit)
        sql = f"""
            SELECT * FROM profile_candidates
            {where}
            ORDER BY updated_at DESC, candidate_id ASC
            LIMIT ?
        """
        with self._connect() as conn:
            rows = conn.execute(sql, params).fetchall()
        return [self._row_to_dict(row) for row in rows]

    def get_candidate(self, candidate_id: str) -> Optional[Dict[str, Any]]:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT * FROM profile_candidates WHERE candidate_id = ?",
                (candidate_id,),
            ).fetchone()
        return self._row_to_dict(row) if row else None

    def mark_promoted(self, candidate_id: str, promoted_record_id: str, promoted_import_id: str) -> None:
        now = datetime.now().isoformat(timespec="seconds")
        with self._connect() as conn:
            conn.execute(
                """
                UPDATE profile_candidates
                SET promotion_status = 'promoted',
                    promoted_record_id = ?,
                    promoted_import_id = ?,
                    updated_at = ?
                WHERE candidate_id = ?
                """,
                (promoted_record_id, promoted_import_id, now, candidate_id),
            )

    def get_statistics(self) -> Dict[str, Any]:
        with self._connect() as conn:
            total = conn.execute("SELECT COUNT(*) FROM profile_candidates").fetchone()[0]
            by_quality = dict(
                conn.execute(
                    """
                    SELECT COALESCE(quality_decision, '未知'), COUNT(*)
                    FROM profile_candidates
                    GROUP BY quality_decision
                    """
                ).fetchall()
            )
            by_promotion = dict(
                conn.execute(
                    """
                    SELECT COALESCE(promotion_status, '未知'), COUNT(*)
                    FROM profile_candidates
                    GROUP BY promotion_status
                    """
                ).fetchall()
            )
        return {
            "total_candidates": total,
            "quality_decision_stats": by_quality,
            "promotion_status_stats": by_promotion,
        }

    def _upsert_candidate(
        self,
        conn: sqlite3.Connection,
        candidate: Dict[str, Any],
        batch_id: str,
        now: str,
    ) -> None:
        source = candidate.get("source", {})
        assistant = candidate.get("assistant_record", {})
        candidate_profile = candidate.get("candidate_profile", {})
        quality = candidate.get("quality_control", {})

        conn.execute(
            """
            INSERT INTO profile_candidates (
                candidate_id, latest_batch_id, source_file, source_file_sha256,
                sheet_name, excel_row_number, work_date, assistant_name,
                lawyer_name, matter_code, work_type, task_source, assigned_task,
                completed_work, deliverable_type, duration_minutes, current_status,
                blocker_reason, lawyer_feedback, next_follow_up, habit_suitability,
                note, candidate_rule, quality_decision, sensitive_flags_json,
                sensitive_flags_by_field_json, candidate_json, review_status,
                promotion_status, promoted_record_id, promoted_import_id,
                created_at, updated_at
            ) VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )
            ON CONFLICT(candidate_id) DO UPDATE SET
                latest_batch_id = excluded.latest_batch_id,
                source_file = excluded.source_file,
                source_file_sha256 = excluded.source_file_sha256,
                sheet_name = excluded.sheet_name,
                excel_row_number = excluded.excel_row_number,
                work_date = excluded.work_date,
                assistant_name = excluded.assistant_name,
                lawyer_name = excluded.lawyer_name,
                matter_code = excluded.matter_code,
                work_type = excluded.work_type,
                task_source = excluded.task_source,
                assigned_task = excluded.assigned_task,
                completed_work = excluded.completed_work,
                deliverable_type = excluded.deliverable_type,
                duration_minutes = excluded.duration_minutes,
                current_status = excluded.current_status,
                blocker_reason = excluded.blocker_reason,
                lawyer_feedback = excluded.lawyer_feedback,
                next_follow_up = excluded.next_follow_up,
                habit_suitability = excluded.habit_suitability,
                note = excluded.note,
                candidate_rule = excluded.candidate_rule,
                quality_decision = excluded.quality_decision,
                sensitive_flags_json = excluded.sensitive_flags_json,
                sensitive_flags_by_field_json = excluded.sensitive_flags_by_field_json,
                candidate_json = excluded.candidate_json,
                review_status = CASE
                    WHEN profile_candidates.promotion_status = 'promoted'
                    THEN profile_candidates.review_status
                    ELSE excluded.review_status
                END,
                promotion_status = CASE
                    WHEN profile_candidates.promotion_status = 'promoted'
                    THEN profile_candidates.promotion_status
                    ELSE excluded.promotion_status
                END,
                promoted_record_id = profile_candidates.promoted_record_id,
                promoted_import_id = profile_candidates.promoted_import_id,
                updated_at = excluded.updated_at
            """,
            (
                candidate["candidate_id"],
                batch_id,
                source.get("source_file"),
                source.get("source_file_sha256"),
                source.get("sheet_name"),
                source.get("excel_row_number"),
                assistant.get("work_date"),
                assistant.get("assistant_name"),
                assistant.get("lawyer_name"),
                assistant.get("matter_code"),
                assistant.get("work_type"),
                assistant.get("task_source"),
                assistant.get("assigned_task"),
                assistant.get("completed_work"),
                assistant.get("deliverable_type"),
                assistant.get("duration_minutes"),
                assistant.get("current_status"),
                assistant.get("blocker_reason"),
                assistant.get("lawyer_feedback"),
                assistant.get("next_follow_up"),
                assistant.get("habit_suitability"),
                assistant.get("note"),
                candidate_profile.get("candidate_rule"),
                quality.get("privacy_decision"),
                json.dumps(quality.get("sensitive_flags", []), ensure_ascii=False),
                json.dumps(quality.get("sensitive_flags_by_field", {}), ensure_ascii=False),
                json.dumps(candidate, ensure_ascii=False),
                candidate.get("review_status", "pending_review"),
                candidate.get("promotion_status", "not_promoted"),
                candidate.get("promoted_record_id"),
                candidate.get("promoted_import_id"),
                now,
                now,
            ),
        )

    def _row_to_dict(self, row: sqlite3.Row) -> Dict[str, Any]:
        data = dict(row)
        for field in ["sensitive_flags_json", "sensitive_flags_by_field_json", "candidate_json"]:
            if field in data:
                data[field.replace("_json", "")] = json.loads(data[field])
                del data[field]
        return data

    def _build_batch_id(self, meta: Dict[str, Any], path: Path) -> str:
        raw = "|".join(
            [
                meta.get("schema_version", ""),
                meta.get("generated_at", ""),
                meta.get("input_file_sha256", ""),
                str(path),
            ]
        )
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:20]


def _is_assistant_row_suitable(value: str, include_uncertain: bool) -> bool:
    text = value.strip()
    if text.startswith("是"):
        return True
    if include_uncertain and "不确定" in text:
        return True
    return False


def build_candidate_rule(payload: Dict[str, str]) -> str:
    work_type = payload.get("work_type") or "协作"
    feedback = payload.get("lawyer_feedback") or "待补充律师反馈"
    deliverable = payload.get("deliverable_type") or "相应交付物"
    blocker = payload.get("blocker_reason") or ""

    rule = f"在【{work_type}】任务中，律师偏好【{feedback}】；常用交付物为【{deliverable}】。"
    if blocker and blocker != "无明显卡点":
        rule += f" 如遇【{blocker}】，需提前提示并补齐条件。"
    return rule


def clean_assistant_workbook_to_candidates(
    input_path: Path,
    output_dir: Path,
    config: Dict[str, Any],
    requested_sheet: Optional[str] = None,
    candidate_db_path: Optional[str] = None,
) -> Tuple[Path, CandidateIngestionSummary]:
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")
    output_dir.mkdir(parents=True, exist_ok=True)

    assistant_config = config.get("assistant_candidates", {})
    wb = load_workbook(input_path, data_only=True, read_only=True)
    sheet_name = choose_assistant_sheet(
        wb,
        requested_sheet,
        assistant_config.get("sheet_candidates", ["每日协作记录"]),
    )
    ws = wb[sheet_name]

    header_row, header_map = _find_header_row(
        ws,
        ASSISTANT_HEADER_SYNONYMS,
        min_score=6,
        required_fields=ASSISTANT_REQUIRED_HEADERS,
    )

    source_hash = file_sha256(input_path)
    include_uncertain = bool(assistant_config.get("include_uncertain", False))
    default_ingestion_level = assistant_config.get("default_ingestion_level", "B可用样本")

    candidates: List[Dict[str, Any]] = []
    stats = {
        "total_excel_rows_scanned": 0,
        "empty_rows_skipped": 0,
        "rows_skipped_not_suitable": 0,
        "candidates_exported": 0,
        "candidates_need_manual_review": 0,
        "candidates_low_value": 0,
        "sensitive_flag_counts": {},
    }

    for row_idx in range(header_row + 1, ws.max_row + 1):
        stats["total_excel_rows_scanned"] += 1
        payload: Dict[str, str] = {}
        sensitive_flags_by_field: Dict[str, List[str]] = {}
        all_flags: List[str] = []

        for field in ASSISTANT_FIELD_ORDER:
            col_idx = header_map.get(field)
            raw_value = ws.cell(row_idx, col_idx).value if col_idx else ""
            if field == "work_date":
                sanitized, flags = stringify_cell(raw_value), []
            else:
                sanitized, flags = sanitize_text(raw_value)
            payload[field] = sanitized
            if flags:
                sensitive_flags_by_field[field] = flags
                all_flags.extend(flags)

        if not any(payload.values()):
            stats["empty_rows_skipped"] += 1
            continue

        if not _is_assistant_row_suitable(payload.get("habit_suitability", ""), include_uncertain):
            stats["rows_skipped_not_suitable"] += 1
            continue

        missing_core = []
        if not payload.get("lawyer_feedback"):
            missing_core.append("lawyer_feedback")

        all_flags_unique = sorted(set(all_flags))
        for flag in all_flags_unique:
            stats["sensitive_flag_counts"][flag] = stats["sensitive_flag_counts"].get(flag, 0) + 1

        privacy_decision = "pass"
        if all_flags_unique:
            privacy_decision = "needs_manual_review"
            stats["candidates_need_manual_review"] += 1
        if missing_core:
            privacy_decision = "low_value_or_incomplete"
            stats["candidates_low_value"] += 1

        candidate_rule = build_candidate_rule(payload)
        candidate_id = stable_candidate_id(input_path.name, sheet_name, row_idx, payload)
        candidate = {
            "candidate_id": candidate_id,
            "source": {
                "source_file": input_path.name,
                "source_file_sha256": source_hash,
                "sheet_name": sheet_name,
                "excel_row_number": row_idx,
            },
            "assistant_record": payload,
            "candidate_profile": {
                "data_source": "助理协作记录",
                "role_pattern": "律师-助理协作",
                "work_type": payload.get("work_type", ""),
                "lawyer_feedback": payload.get("lawyer_feedback", ""),
                "candidate_rule": candidate_rule,
                "default_ingestion_level": default_ingestion_level,
                "review_tag": "协作偏好",
                "profile_update_action": "新增规则",
            },
            "quality_control": {
                "missing_core_fields": missing_core,
                "privacy_decision": privacy_decision,
                "sensitive_flags": all_flags_unique,
                "sensitive_flags_by_field": sensitive_flags_by_field,
            },
            "review_status": "pending_review",
            "promotion_status": "not_promoted",
        }
        candidates.append(candidate)

    stats["candidates_exported"] = len(candidates)

    exported = {
        "meta": {
            "schema_version": assistant_config.get("schema_version", "lawyer_profile_candidate_v1"),
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "input_file": input_path.name,
            "input_file_sha256": source_hash,
            "sheet_name": sheet_name,
            "header_row": header_row,
            "privacy_principle": "仅输出脱敏后的协作偏好候选；候选需批处理确认后才可升格入正式画像库。",
        },
        "stats": stats,
        "candidates": candidates,
    }

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{_safe_output_name(input_path.stem)}_candidates_{timestamp}.json"
    pretty = bool(config.get("output", {}).get("pretty_json", True))
    _write_json(exported, output_path, pretty=pretty)

    store = ProfileCandidateStore(candidate_db_path or config.get("candidate_db_path", "data/profile_candidates.db"))
    summary = store.ingest_candidate_json_file(str(output_path))
    return output_path, summary


def candidate_to_profile_record(candidate: Dict[str, Any]) -> Dict[str, Any]:
    assistant = candidate.get("assistant_record") or candidate.get("candidate", {}).get("assistant_record") or {}
    candidate_json = candidate.get("candidate") if "candidate" in candidate else candidate
    source = candidate_json.get("source", {})
    profile = candidate_json.get("candidate_profile", {})
    quality = candidate_json.get("quality_control", {})
    candidate_id = candidate_json.get("candidate_id") or candidate.get("candidate_id")
    candidate_rule = profile.get("candidate_rule") or candidate.get("candidate_rule") or ""
    record_id = stable_promoted_record_id(candidate_id, candidate_rule)

    return {
        "record_id": record_id,
        "source": {
            "source_file": source.get("source_file"),
            "source_file_sha256": source.get("source_file_sha256"),
            "sheet_name": source.get("sheet_name"),
            "excel_row_number": source.get("excel_row_number"),
        },
        "taxonomy": {
            "data_source": "助理协作记录",
            "matter_type": "协作流程",
            "stage": assistant.get("work_type", ""),
            "representativeness": assistant.get("habit_suitability", ""),
        },
        "judgment_model": {
            "conflict_structure": assistant.get("blocker_reason") or "协作偏好",
            "role_pattern": "律师-助理协作",
            "client_goal": "提高协作效率",
            "key_constraints": assistant.get("blocker_reason", ""),
            "first_judgment": "可沉淀协作习惯",
            "abstract_reason": assistant.get("completed_work") or assistant.get("assigned_task", ""),
            "strategy_choice": assistant.get("lawyer_feedback", ""),
            "value_order": "效率与质量优先",
            "risk_communication": "",
            "handling_temperature": "",
            "reusable_rule": candidate_rule,
        },
        "review_and_ingestion": {
            "collection_date": assistant.get("work_date", ""),
            "collection_id_or_masked_ref": candidate_id,
            "result_review": assistant.get("current_status", ""),
            "note": "；".join(
                part for part in [assistant.get("next_follow_up", ""), assistant.get("note", "")]
                if part
            ),
            "desensitization_status": "已脱敏",
            "ingestion_level": profile.get("default_ingestion_level", "B可用样本"),
            "cleaning_note": "由助理协作记录候选池升格",
            "review_tag": profile.get("review_tag", "协作偏好"),
            "profile_update_action": profile.get("profile_update_action", "新增规则"),
        },
        "quality_control": {
            "missing_core_fields": quality.get("missing_core_fields", []),
            "privacy_decision": "pass",
            "sensitive_flags": [],
            "sensitive_flags_by_field": {},
        },
    }


def promote_candidates(
    candidate_db_path: str,
    client_profile_db_path: str,
    output_dir: Path,
    candidate_ids: Optional[Sequence[str]] = None,
    all_pass: bool = False,
    limit: int = 100,
) -> CandidatePromotionSummary:
    store = ProfileCandidateStore(candidate_db_path)
    if all_pass:
        candidates = store.list_candidates(
            promotion_status="not_promoted",
            quality_decision="pass",
            limit=limit,
        )
    else:
        if not candidate_ids:
            raise ValueError("请提供 candidate_ids，或使用 all_pass=True。")
        candidates = []
        for candidate_id in candidate_ids:
            candidate = store.get_candidate(candidate_id)
            if candidate:
                candidates.append(candidate)

    records = []
    promoted_pairs: List[Tuple[str, str]] = []
    skipped = 0
    for candidate in candidates:
        candidate_payload = candidate.get("candidate", {})
        quality = candidate_payload.get("quality_control", {})
        if candidate.get("promotion_status") == "promoted":
            skipped += 1
            continue
        if candidate.get("quality_decision") != "pass" or quality.get("privacy_decision") != "pass":
            skipped += 1
            continue
        record = candidate_to_profile_record(candidate_payload)
        records.append(record)
        promoted_pairs.append((candidate["candidate_id"], record["record_id"]))

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    exported = {
        "meta": {
            "schema_version": "lawyer_profile_memory_v1",
            "generated_at": now,
            "input_file": "profile_candidates.db",
            "input_file_sha256": "",
            "sheet_name": "profile_candidates",
            "header_row": 0,
            "privacy_principle": "由已通过质量检查的助理协作候选升格；不包含原始案件细节。",
        },
        "stats": {
            "records_exported": len(records),
            "records_promoted_from_candidates": len(records),
            "records_skipped": skipped,
        },
        "records": records,
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"assistant_candidates_promoted_{timestamp}.json"
    _write_json(exported, output_path, pretty=True)

    profile_store = ClientProfileStore(client_profile_db_path)
    ingestion = profile_store.ingest_json_file(str(output_path))
    for candidate_id, record_id in promoted_pairs:
        store.mark_promoted(candidate_id, record_id, ingestion.import_id)

    return CandidatePromotionSummary(
        output_file=str(output_path),
        records_seen=len(candidates),
        records_promoted=ingestion.records_upserted,
        records_skipped=skipped + ingestion.records_skipped,
        import_id=ingestion.import_id,
        client_profile_db_path=client_profile_db_path,
    )


# ============================================================
# 7. CLI
# ============================================================

def _print_json(data: Dict[str, Any]) -> None:
    print(json.dumps(data, ensure_ascii=False, indent=2))


def _parse_candidate_ids(value: str) -> List[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="律师画像数据批处理管线")
    subparsers = parser.add_subparsers(dest="command")

    clean_profile = subparsers.add_parser("clean-profile", help="清洗用户版/开发版采集表")
    clean_profile.add_argument("--input", required=True, help="输入 Excel 文件路径")
    clean_profile.add_argument("--output-dir", default="output", help="输出文件夹")
    clean_profile.add_argument("--config", default="config.json", help="配置文件")
    clean_profile.add_argument("--sheet", default="", help="可选：指定工作表名")
    clean_profile.add_argument("--use-model", default="false", help="是否启用模型占位逻辑")
    clean_profile.add_argument("--import-db", default="", help="可选：清洗后导入正式画像 SQLite")

    clean_assistant = subparsers.add_parser("clean-assistant", help="清洗助理协作表到候选画像池")
    clean_assistant.add_argument("--input", required=True, help="输入 Excel 文件路径")
    clean_assistant.add_argument("--output-dir", default="output", help="输出文件夹")
    clean_assistant.add_argument("--config", default="config.json", help="配置文件")
    clean_assistant.add_argument("--sheet", default="", help="可选：指定工作表名")
    clean_assistant.add_argument("--candidate-db", default="", help="候选池 SQLite 路径")

    promote = subparsers.add_parser("promote-candidates", help="从候选画像池升格到正式画像库")
    promote.add_argument("--config", default="config.json", help="配置文件")
    promote.add_argument("--candidate-db", default="", help="候选池 SQLite 路径")
    promote.add_argument("--client-profile-db", default="", help="正式画像 SQLite 路径")
    promote.add_argument("--output-dir", default="output", help="输出文件夹")
    promote.add_argument("--candidate-ids", default="", help="逗号分隔的 candidate_id 列表")
    promote.add_argument("--all-pass", action="store_true", help="升格全部未升格且质量通过的候选")
    promote.add_argument("--limit", type=int, default=100, help="all-pass 最大升格数量")

    list_candidates = subparsers.add_parser("list-candidates", help="列出候选画像")
    list_candidates.add_argument("--config", default="config.json", help="配置文件")
    list_candidates.add_argument("--candidate-db", default="", help="候选池 SQLite 路径")
    list_candidates.add_argument("--limit", type=int, default=20)
    list_candidates.add_argument("--quality-decision", default="")
    list_candidates.add_argument("--promotion-status", default="not_promoted")

    stats = subparsers.add_parser("candidate-stats", help="查看候选池统计")
    stats.add_argument("--config", default="config.json", help="配置文件")
    stats.add_argument("--candidate-db", default="", help="候选池 SQLite 路径")

    args = parser.parse_args(list(argv) if argv is not None else None)
    if not args.command:
        parser.print_help()
        return 1

    config = load_pipeline_config(Path(getattr(args, "config", "")) if getattr(args, "config", "") else None)

    if args.command == "clean-profile":
        use_model = parse_bool(args.use_model) or bool(config.get("model", {}).get("use_model", False))
        output_path = clean_profile_workbook(
            input_path=Path(args.input),
            output_dir=Path(args.output_dir),
            config=config,
            requested_sheet=args.sheet.strip() or None,
            use_model=use_model,
        )
        result: Dict[str, Any] = {"output_file": str(output_path)}
        if args.import_db:
            summary = ClientProfileStore(args.import_db).ingest_json_file(str(output_path))
            result["import"] = {
                "import_id": summary.import_id,
                "records_seen": summary.records_seen,
                "records_upserted": summary.records_upserted,
                "records_skipped": summary.records_skipped,
                "db_path": args.import_db,
            }
        print(f"[OK] 已生成：{output_path}")
        if result.get("import"):
            _print_json(result["import"])
        return 0

    if args.command == "clean-assistant":
        candidate_db = args.candidate_db or config.get("candidate_db_path", "data/profile_candidates.db")
        output_path, summary = clean_assistant_workbook_to_candidates(
            input_path=Path(args.input),
            output_dir=Path(args.output_dir),
            config=config,
            requested_sheet=args.sheet.strip() or None,
            candidate_db_path=candidate_db,
        )
        print(f"[OK] 已生成候选画像：{output_path}")
        _print_json(
            {
                "batch_id": summary.batch_id,
                "candidates_seen": summary.candidates_seen,
                "candidates_upserted": summary.candidates_upserted,
                "db_path": summary.db_path,
            }
        )
        return 0

    if args.command == "promote-candidates":
        candidate_db = args.candidate_db or config.get("candidate_db_path", "data/profile_candidates.db")
        client_profile_db = args.client_profile_db or config.get("client_profile_db_path", "data/client_profiles.db")
        summary = promote_candidates(
            candidate_db_path=candidate_db,
            client_profile_db_path=client_profile_db,
            output_dir=Path(args.output_dir),
            candidate_ids=_parse_candidate_ids(args.candidate_ids),
            all_pass=args.all_pass,
            limit=args.limit,
        )
        _print_json(summary.__dict__)
        return 0

    if args.command == "list-candidates":
        candidate_db = args.candidate_db or config.get("candidate_db_path", "data/profile_candidates.db")
        store = ProfileCandidateStore(candidate_db)
        candidates = store.list_candidates(
            quality_decision=args.quality_decision or None,
            promotion_status=args.promotion_status or None,
            limit=args.limit,
        )
        for candidate in candidates:
            print(
                f"{candidate['candidate_id']} | {candidate.get('quality_decision')} | "
                f"{candidate.get('promotion_status')} | {candidate.get('work_type')} | "
                f"{candidate.get('lawyer_feedback')}"
            )
        if not candidates:
            print("[INFO] No candidates found.")
        return 0

    if args.command == "candidate-stats":
        candidate_db = args.candidate_db or config.get("candidate_db_path", "data/profile_candidates.db")
        store = ProfileCandidateStore(candidate_db)
        _print_json(store.get_statistics())
        return 0

    return 1


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(1)
