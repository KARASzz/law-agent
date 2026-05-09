# -*- coding: utf-8 -*-
"""
律师客户画像采集表清洗入库脚本

目标：
- 从 input 文件夹中的 Excel 采集表读取数据；
- 只抽取“抽象行为逻辑”字段，不抽取案件细节；
- 对文本做基础脱敏扫描与替换；
- 输出结构化 JSON，供后续轻量文本画像库 / 阿里云记忆库导入前使用。

注意：
- 本脚本默认不调用模型；
- 如需模型辅助抽象、归一化、规则改写，请在 normalize_with_model() 内补充你的模型调用代码；
- 模型调用前请确保只发送已脱敏字段。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ============================================================
# 1. 字段映射：兼容“客户版”和“开发版”
# ============================================================

HEADER_SYNONYMS: Dict[str, List[str]] = {
    "collection_date": ["采集日期", "采集日期*"],
    "collection_id": ["采集ID", "脱敏案号"],
    "data_source": ["数据来源", "数据来源*"],
    "matter_type": ["案件类型", "案件类型*", "业务类型"],
    "stage": ["阶段", "阶段*", "案件阶段"],
    "conflict_structure": ["冲突结构", "冲突结构*"],
    "role_pattern": ["角色格局"],
    "client_goal": ["客户核心诉求"],
    "key_constraints": ["关键约束"],
    "first_judgment": ["第一判断", "第一判断*"],
    "abstract_reason": ["判断理由_抽象", "判断理由"],
    "strategy_choice": ["策略选择", "策略选择*"],
    "value_order": ["价值排序"],
    "risk_communication": ["风险沟通方式"],
    "handling_temperature": ["处理温度"],
    "result_review": ["结果/复盘", "复盘一句话_可空"],
    "reusable_rule": ["可复用规则"],
    "representativeness": ["代表性"],
    "note": ["备注_不写细节", "备注"],
    # 以下通常只存在于开发版，不应要求客户填写。
    "desensitization_status": ["脱敏状态", "脱敏状态*"],
    "ingestion_level": ["入库等级", "入库等级*"],
    "cleaning_note": ["清洗备注"],
    "review_tag": ["复盘标签"],
    "profile_update_action": ["画像更新动作"],
}

CORE_FIELDS = [
    "conflict_structure",
    "first_judgment",
    "strategy_choice",
    "reusable_rule",
]

# 只输出这些抽象字段；不输出整行原始数据。
OUTPUT_FIELD_ORDER = [
    "collection_date",
    "collection_id",
    "data_source",
    "matter_type",
    "stage",
    "conflict_structure",
    "role_pattern",
    "client_goal",
    "key_constraints",
    "first_judgment",
    "abstract_reason",
    "strategy_choice",
    "value_order",
    "risk_communication",
    "handling_temperature",
    "result_review",
    "reusable_rule",
    "representativeness",
    "note",
    "desensitization_status",
    "ingestion_level",
    "cleaning_note",
    "review_tag",
    "profile_update_action",
]


# ============================================================
# 2. 基础脱敏规则：保守处理，宁可标记复核，不替你做事实判断
# ============================================================

SENSITIVE_PATTERNS: List[Tuple[str, re.Pattern]] = [
    ("身份证号", re.compile(r"(?<!\d)\d{17}[\dXx](?!\d)")),
    ("手机号", re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)")),
    ("邮箱", re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")),
    ("疑似案号", re.compile(r"[（(]\d{4}[）)][^，。；;\s]{2,40}?号")),
    ("疑似金额", re.compile(r"\d+(?:\.\d+)?\s*(?:万|万元|元|亿|亿元)")),
    ("疑似具体日期", re.compile(r"\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}日?")),
]

# 机构名不自动替换，避免误伤“法院倾向/仲裁倾向”这类抽象表达；只做风险提示。
SENSITIVE_FLAG_ONLY: List[Tuple[str, re.Pattern]] = [
    ("疑似具体司法/行政机构", re.compile(r"[^，。；;\s]{2,20}(人民法院|检察院|公安局|仲裁委员会|市场监督管理局)")),
    ("疑似自然人姓名提示", re.compile(r"(张三|李四|王五|某某|当事人姓名|对方姓名|法官姓名|律师姓名)")),
]


def normalize_header(text: Any) -> str:
    if text is None:
        return ""
    return str(text).strip().replace("*", "").replace(" ", "")


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def sanitize_text(value: Any) -> Tuple[str, List[str]]:
    """返回：脱敏后的文本、命中的敏感风险标签。"""
    text = stringify_cell(value)
    if not text:
        return "", []

    flags: List[str] = []
    sanitized = text

    for name, pattern in SENSITIVE_PATTERNS:
        if pattern.search(sanitized):
            flags.append(name)
            sanitized = pattern.sub(f"[已脱敏_{name}]", sanitized)

    for name, pattern in SENSITIVE_FLAG_ONLY:
        if pattern.search(sanitized):
            flags.append(name)

    # 限制异常超长文本进入记忆库，避免夹带案件材料。
    if len(sanitized) > 800:
        flags.append("文本过长_建议人工复核")
        sanitized = sanitized[:800] + "...[已截断_待复核]"

    return sanitized, sorted(set(flags))


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def stable_record_id(source_file: str, sheet_name: str, row_number: int, row_payload: Dict[str, str]) -> str:
    seed = json.dumps({
        "source_file": source_file,
        "sheet_name": sheet_name,
        "row_number": row_number,
        "core": {k: row_payload.get(k, "") for k in CORE_FIELDS},
    }, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(seed.encode("utf-8")).hexdigest()[:20]


# ============================================================
# 3. Excel 读取与表头识别
# ============================================================

def build_synonym_lookup() -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    for canonical, aliases in HEADER_SYNONYMS.items():
        for alias in aliases:
            lookup[normalize_header(alias)] = canonical
    return lookup


def find_header_row(ws, max_scan_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    """在前 max_scan_rows 行内识别表头行，返回 header_row_index 与 canonical_field -> column_index。"""
    synonym_lookup = build_synonym_lookup()
    best_row = -1
    best_score = 0
    best_map: Dict[str, int] = {}

    for row_idx in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_map: Dict[str, int] = {}
        score = 0
        for col_idx in range(1, ws.max_column + 1):
            header = normalize_header(ws.cell(row_idx, col_idx).value)
            if header in synonym_lookup:
                canonical = synonym_lookup[header]
                # 若重复出现，优先保留靠前列。
                row_map.setdefault(canonical, col_idx)
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_map = row_map

    if best_score < 4:
        raise ValueError(
            "未识别到有效表头。请确认工作表为客户版/开发版采集表，且表头没有被大幅改名。"
        )

    return best_row, best_map


def choose_sheet(wb, requested_sheet: Optional[str], sheet_candidates: List[str]) -> str:
    if requested_sheet:
        if requested_sheet not in wb.sheetnames:
            raise ValueError(f"指定工作表不存在：{requested_sheet}；当前工作表：{wb.sheetnames}")
        return requested_sheet

    for name in sheet_candidates:
        if name in wb.sheetnames:
            return name

    # 兜底：找名称含“填写表”或“采集表”的表。
    for name in wb.sheetnames:
        if "填写表" in name or "采集表" in name:
            return name

    raise ValueError(f"没有找到采集表工作表。当前工作表：{wb.sheetnames}")


# ============================================================
# 4. 模型占位符：默认不启用
# ============================================================

def normalize_with_model(record: Dict[str, Any], model_config: Dict[str, Any]) -> Dict[str, Any]:
    """
    可选：用模型进一步把采集记录归一化为画像规则。

    默认直接返回原记录，不做外部请求。

    你可以在这里接入阿里云或其他模型：
    - 只发送 record["judgment_model"] 里的脱敏抽象字段；
    - 不发送客户姓名、案号、法院、对方身份、金额、证据原文；
    - 要求模型只返回 JSON；
    - 对模型输出再做一次 JSON schema 校验。

    伪代码：
    api_key = os.environ.get(model_config["api_key_env"])
    endpoint = model_config["endpoint"]
    prompt = build_prompt(record)
    response = requests.post(endpoint, headers=..., json=...)
    record["model_normalized"] = response.json()
    return record
    """
    record["model_normalized"] = None
    record["model_note"] = "model disabled; fill normalize_with_model() when needed"
    return record


# ============================================================
# 5. 清洗主流程
# ============================================================

def parse_bool(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "启用"}


def load_config(path: Optional[Path]) -> Dict[str, Any]:
    default = {
        "schema_version": "lawyer_profile_memory_v1",
        "sheet_candidates": ["01_用户填写表", "01_极简采集表"],
        "output": {"pretty_json": True, "also_write_jsonl": False},
        "model": {"use_model": False},
    }
    if not path:
        return default
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as f:
        loaded = json.load(f)
    # 浅合并即可。
    for k, v in loaded.items():
        if isinstance(v, dict) and isinstance(default.get(k), dict):
            default[k].update(v)
        else:
            default[k] = v
    return default


def clean_workbook(input_path: Path, output_dir: Path, config: Dict[str, Any], requested_sheet: Optional[str], use_model: bool) -> Path:
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, data_only=True, read_only=False)
    sheet_name = choose_sheet(wb, requested_sheet, config.get("sheet_candidates", []))
    ws = wb[sheet_name]

    header_row, header_map = find_header_row(ws)

    records: List[Dict[str, Any]] = []
    stats = {
        "total_excel_rows_scanned": 0,
        "empty_rows_skipped": 0,
        "records_exported": 0,
        "records_need_manual_review": 0,
        "sensitive_flag_counts": {},
        "missing_core_field_counts": {field: 0 for field in CORE_FIELDS},
    }

    source_hash = file_sha256(input_path)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        stats["total_excel_rows_scanned"] += 1

        raw_payload: Dict[str, str] = {}
        sensitive_flags_by_field: Dict[str, List[str]] = {}
        all_flags: List[str] = []

        for field in OUTPUT_FIELD_ORDER:
            col_idx = header_map.get(field)
            raw_value = ws.cell(row_idx, col_idx).value if col_idx else ""
            if field == "collection_date":
                sanitized, flags = stringify_cell(raw_value), []
            else:
                sanitized, flags = sanitize_text(raw_value)
            raw_payload[field] = sanitized
            if flags:
                sensitive_flags_by_field[field] = flags
                all_flags.extend(flags)

        # 跳过全空行。
        if not any(raw_payload.values()):
            stats["empty_rows_skipped"] += 1
            continue

        missing_core = [field for field in CORE_FIELDS if not raw_payload.get(field)]
        for field in missing_core:
            stats["missing_core_field_counts"][field] += 1

        all_flags_unique = sorted(set(all_flags))
        for flag in all_flags_unique:
            stats["sensitive_flag_counts"][flag] = stats["sensitive_flag_counts"].get(flag, 0) + 1

        privacy_decision = "pass"
        if all_flags_unique:
            privacy_decision = "needs_manual_review"
        if len(missing_core) >= 3:
            privacy_decision = "low_value_or_incomplete"
        if privacy_decision == "needs_manual_review":
            stats["records_need_manual_review"] += 1

        record = {
            "record_id": stable_record_id(input_path.name, sheet_name, row_idx, raw_payload),
            "source": {
                "source_file": input_path.name,
                "source_file_sha256": source_hash,
                "sheet_name": sheet_name,
                "excel_row_number": row_idx,
            },
            "taxonomy": {
                "data_source": raw_payload.get("data_source", ""),
                "matter_type": raw_payload.get("matter_type", ""),
                "stage": raw_payload.get("stage", ""),
                "representativeness": raw_payload.get("representativeness", ""),
            },
            "judgment_model": {
                "conflict_structure": raw_payload.get("conflict_structure", ""),
                "role_pattern": raw_payload.get("role_pattern", ""),
                "client_goal": raw_payload.get("client_goal", ""),
                "key_constraints": raw_payload.get("key_constraints", ""),
                "first_judgment": raw_payload.get("first_judgment", ""),
                "abstract_reason": raw_payload.get("abstract_reason", ""),
                "strategy_choice": raw_payload.get("strategy_choice", ""),
                "value_order": raw_payload.get("value_order", ""),
                "risk_communication": raw_payload.get("risk_communication", ""),
                "handling_temperature": raw_payload.get("handling_temperature", ""),
                "reusable_rule": raw_payload.get("reusable_rule", ""),
            },
            "review_and_ingestion": {
                "collection_date": raw_payload.get("collection_date", ""),
                "collection_id_or_masked_ref": raw_payload.get("collection_id", ""),
                "result_review": raw_payload.get("result_review", ""),
                "note": raw_payload.get("note", ""),
                "desensitization_status": raw_payload.get("desensitization_status", ""),
                "ingestion_level": raw_payload.get("ingestion_level", ""),
                "cleaning_note": raw_payload.get("cleaning_note", ""),
                "review_tag": raw_payload.get("review_tag", ""),
                "profile_update_action": raw_payload.get("profile_update_action", ""),
            },
            "quality_control": {
                "missing_core_fields": missing_core,
                "privacy_decision": privacy_decision,
                "sensitive_flags": all_flags_unique,
                "sensitive_flags_by_field": sensitive_flags_by_field,
            },
        }

        if use_model:
            record = normalize_with_model(record, config.get("model", {}))

        records.append(record)

    stats["records_exported"] = len(records)

    exported = {
        "meta": {
            "schema_version": config.get("schema_version", "lawyer_profile_memory_v1"),
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "input_file": input_path.name,
            "input_file_sha256": source_hash,
            "sheet_name": sheet_name,
            "header_row": header_row,
            "privacy_principle": "仅输出脱敏后的抽象行为逻辑字段；不输出整行原始案件信息。",
        },
        "stats": stats,
        "records": records,
    }

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_base = re.sub(r"[^0-9A-Za-z_\-\u4e00-\u9fa5]", "_", input_path.stem)
    output_path = output_dir / f"{safe_base}_cleaned_{timestamp}.json"

    pretty = bool(config.get("output", {}).get("pretty_json", True))
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(exported, f, ensure_ascii=False, indent=2 if pretty else None)

    if bool(config.get("output", {}).get("also_write_jsonl", False)):
        jsonl_path = output_dir / f"{safe_base}_records_{timestamp}.jsonl"
        with jsonl_path.open("w", encoding="utf-8") as f:
            for rec in records:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="律师客户画像采集表清洗为 JSON")
    parser.add_argument("--input", required=True, help="输入 Excel 文件路径，例如 input\\本周采集表.xlsx")
    parser.add_argument("--output-dir", default="output", help="输出文件夹，默认 output")
    parser.add_argument("--config", default="config.json", help="配置文件，默认 config.json")
    parser.add_argument("--sheet", default="", help="可选：指定工作表名，不填则自动识别")
    parser.add_argument("--use-model", default="false", help="是否启用模型占位逻辑：true/false，默认 false")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    config = load_config(Path(args.config) if args.config else None)

    # 命令行优先级高于 config。
    use_model = parse_bool(args.use_model) or bool(config.get("model", {}).get("use_model", False))

    output_path = clean_workbook(
        input_path=input_path,
        output_dir=output_dir,
        config=config,
        requested_sheet=args.sheet.strip() or None,
        use_model=use_model,
    )

    print(f"[OK] 已生成：{output_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(1)
